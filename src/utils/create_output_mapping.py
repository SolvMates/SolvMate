import os
from pathlib import Path
from supabase import create_client, Client
import pandas as pd
from openpyxl import load_workbook
from dotenv import load_dotenv
from get_Value import get_value
import numpy as np
from datetime import datetime, timezone
import re

# Load environment variables for Supabase credentials
load_dotenv()
supabase: Client = create_client(
    os.environ.get("SUPABASE_URL"), os.environ.get("SUPABASE_KEY")
)


def check_cell_value_format(cell_value: str) -> bool:
    """
    Checks if the cell value matches one of the required formats.
    Returns True if the value matches, otherwise False.
    """
    pattern1 = r"^Z-Z\d{4}$"          # e.g., Z-Z0010
    pattern2 = r"^R\d{4}-C\d{4}$"     # e.g., R****-C****
    pattern3 = r"^\d{2}R\d{3}-C\d{4}$" # e.g., 33R***-C****
    pattern4 = r"^E\d{4}-C\d{4}$"     # e.g., E4010-C1370

    if (
        re.match(pattern1, cell_value)
        or re.match(pattern2, cell_value)
        or re.match(pattern3, cell_value)
        or re.match(pattern4, cell_value)
    ):
        return True
    return False


def read_excel_cells(file_path: str) -> pd.DataFrame:
    """
    Reads all cells from the Excel file at file_path and extracts those
    matching the required format. Returns a DataFrame with the results.

    Skips sheets starting with '26' (e.g., '26...').
    """
    # Create an empty DataFrame for the results
    results = pd.DataFrame(
        columns=["QRT_ID", "QRT_NAME", "RC_CODE", "CELL_REFERENCE", "DATA_ID", "ID", "TYPE"]
    )

    # Load Excel file
    excel_file = pd.ExcelFile(file_path)

    # Iterate over all worksheets in the Excel file
    for sheet_name in excel_file.sheet_names:
        if sheet_name.startswith("26"):
            print(f"Skipping sheet: {sheet_name}")  # Debug output
            continue
        df_sheet = excel_file.parse(sheet_name)
        qrt_id = sheet_name[:10].replace(".", "")
        # Iterate over all cells in the DataFrame
        for row in range(df_sheet.shape[0]):  # Rows
            for col in range(df_sheet.shape[1]):  # Columns
                cell_value = df_sheet.iat[row, col]  # Cell content

                # Check if the cell content matches the target format
                if check_cell_value_format(str(cell_value)):
                    # Cell position in Excel format (e.g., A1, B2, ...)
                    cell_position = (
                        f"{chr(65 + col)}{row + 2}"  
                    )
                    id = f"{sheet_name}_{cell_value}"
                    # Add the information to the results DataFrame
                    new_row = pd.DataFrame(
                        {
                            "QRT_ID": [qrt_id],
                            "QRT_NAME": [sheet_name],
                            "RC_CODE": [cell_value],
                            "CELL_REFERENCE": [cell_position],
                            "DATA_ID": None,
                            "ID": [id],
                            "TYPE": None,
                        }
                    )
                    results = pd.concat([results, new_row], ignore_index=True)

    return results

def find_and_add_type_of_cell(dataframe: pd.DataFrame, file_path: str) -> pd.DataFrame:
    workbook = load_workbook(Path(file_path))
    
  
    for index, row in dataframe.iterrows():
        sheet_name = row["QRT_NAME"]
        cell_reference = row["CELL_REFERENCE"]
        
        
        if sheet_name in workbook.sheetnames:
            df_sheet = workbook[sheet_name]
            cell_value = df_sheet[cell_reference].value  
            if cell_value is None:
                cell_value = "None"
            dataframe.at[index, "TYPE"] = cell_value 

    return dataframe

def upsert_data_to_supabase(
    dataframe: pd.DataFrame, table_name: str, unique_column: str
):
    """
    Inserts or updates each row of the DataFrame into the specified Supabase table.
    Uses the unique_column to check for existing records.
    """
    for index, row in dataframe.iterrows():
        unique_value = row[unique_column]

        # Check if the record already exists
        existing_record = (
            supabase.table(table_name)
            .select("*")
            .eq(unique_column, unique_value)
            .execute()
        )

        if existing_record.data:
            # Record exists, update it
            update_data = row[["QRT_ID", "QRT_NAME", "RC_CODE", "CELL_REFERENCE", "TYPE"]].to_dict()
            supabase.table(table_name).update(update_data).eq(unique_column, unique_value).execute()
           
            print(f"Updated record with {unique_column}: {unique_value}")
        else:
            # Record does not exist, insert it
            supabase.table(table_name).insert(row.to_dict()).execute()
            print(f"Inserted new record with {unique_column}: {unique_value}")


def create_output_mapping():
    """
    Main function to create the output mapping:
    - Reads the Excel template
    - Upserts the data to Supabase
    - Writes the result to an Excel file
    """
    goal_dataframe = read_excel_cells(
        file_path="/workspaces/SolvMate/templates/Output_ERGO.xlsx",
    )
    upsert_data_to_supabase(goal_dataframe, "output_mapping", "ID")
    output_path = Path("/workspaces/SolvMate/outputs/Output_Mapping.xlsx")
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        goal_dataframe.to_excel(
            writer, sheet_name="Goal DataFrame", index=False, header=True
        )


if __name__ == "__main__":
    # Script entry point for manual execution
    template_dataframe = read_excel_cells(
        file_path="/workspaces/SolvMate/templates/Output_ERGO.xlsx",
    )
    goal_dataframe = find_and_add_type_of_cell(
        template_dataframe,
        file_path="/workspaces/SolvMate/templates/Output_ERGO_Types.xlsx")
    
    upsert_data_to_supabase(goal_dataframe, "output_mapping", "ID")
    output_path = Path("/workspaces/SolvMate/outputs/Output_Mapping.xlsx")
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        goal_dataframe.to_excel(
            writer, sheet_name="Goal DataFrame", index=False, header=True
        )
