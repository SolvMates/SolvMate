import os
from pathlib import Path
from supabase import create_client, Client
import pandas as pd
from openpyxl import load_workbook
from .input import run_Import
from dotenv import load_dotenv
from .get_Value import get_value
import numpy as np
from datetime import datetime, timezone

# Initialize the Supabase client
load_dotenv()
supabase: Client = create_client(
    os.environ.get("SUPABASE_URL"), os.environ.get("SUPABASE_KEY")
)

def normalize_column_names(df: pd.DataFrame) -> pd.DataFrame:
    """
    Ensure the DataFrame has the required columns with correct names.
    
    Args:
        df (pd.DataFrame): Input DataFrame
        
    Returns:
        pd.DataFrame: DataFrame with normalized column names
    """
    # Handle DATA_ID column
    if "DATA_ID" not in df.columns:
        if "NODE_ID" in df.columns:
            df = df.rename(columns={"NODE_ID": "DATA_ID"})
        else:
            raise ValueError("The input DataFrame must contain a column named 'DATA_ID' or 'NODE_ID'.")

    # Handle VALUE column
    if "VALUE" not in df.columns:
        if "Value" in df.columns:
            df = df.rename(columns={"Value": "VALUE"})
        elif "value" in df.columns:
            df = df.rename(columns={"value": "VALUE"})
        else:
            raise ValueError("The input DataFrame must contain a column named 'VALUE', 'Value', or 'value'.")
            
    return df

def fetch_mapping_data() -> pd.DataFrame:
    """
    Fetch output mapping data from Supabase with pagination.
    
    Returns:
        pd.DataFrame: Mapping data from Supabase
    """
    mapping_data = pd.DataFrame()
    offset = 0
    limit = 1000

    while True:
        response = (
            supabase.table("output_mapping")
            .select("*")
            .range(offset, offset + limit - 1)
            .execute()
        )

        temp_df = pd.DataFrame(response.data)
        if temp_df.empty:
            break

        mapping_data = pd.concat([mapping_data, temp_df])
        offset += limit

    return mapping_data

def write_value_to_cell(sheet, cell: str, value) -> None:
    """
    Write a value to a specific cell in the worksheet.
    
    Args:
        sheet: The worksheet object
        cell (str): Cell reference (e.g., 'A1')
        value: Value to write
    """
    if value is None:
        return
        
    if isinstance(value, list):
        sheet[cell] = value[0]
    else:
        try:
            sheet[cell] = float(value)
        except (ValueError, TypeError):
            sheet[cell] = str(value)

def fill_QRT_from_dataframe(
    dataframe: pd.DataFrame, output_dir="/workspaces/SolvMate/outputs"
) -> None:
    """
    Fill Excel templates with values from DataFrame based on mapping.
    
    Args:
        dataframe (pd.DataFrame): DataFrame containing DATA_ID and VALUE pairs
        output_dir (str): Output directory for filled template
    """
    # Normalize column names
    dataframe = normalize_column_names(dataframe)
    
    # Fetch mapping data
    mapping_data = fetch_mapping_data()
    
    # Load template
    template_path = Path("/workspaces/SolvMate/templates/Output_ERGO.xlsx")
    workbook = load_workbook(template_path)
    
    # Process each worksheet
    for sheet_name in workbook.sheetnames:
        if sheet_name.startswith("26"):
            print(f"Skipping sheet: {sheet_name}")
            continue
            
        sheet = workbook[sheet_name]
        worksheet_mapping = mapping_data[mapping_data["QRT_NAME"] == sheet_name]
        
        # Process each mapping entry
        for mapping_id in worksheet_mapping["ID"].values:
            mapping_row = worksheet_mapping[worksheet_mapping["ID"] == mapping_id]
            cell = mapping_row["CELL_REFERENCE"].values[0]
            data_id = mapping_row["DATA_ID"].values[0]
            
            if isinstance(data_id, str):
                data_id = data_id.strip()
                
            # Try to get value from DataFrame
            try:
                if data_id in dataframe["DATA_ID"].values:
                    value = get_value(data_id, dataframe)
                    print(f"Found value for data_id {data_id}: {value}")
                else:
                    value = ""
                    
                write_value_to_cell(sheet, cell, value)
                    
            except Exception as e:
                print(f"Error processing data_id {data_id} for cell {cell}: {str(e)}")
                continue
    
    # Save filled template
    output_path = Path(output_dir) / "Filled_Output_ERGO.xlsx"
    workbook.save(output_path)
    print(f"Template filled and saved to {output_path}")


if __name__ == "__main__":
    dataframe = run_Import("/workspaces/SolvMate/input/02.01_SAS_Input_MarketR.xls")
    fill_QRT_from_dataframe(dataframe)
